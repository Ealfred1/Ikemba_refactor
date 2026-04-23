"""
Fetches ALL products from the Chowdeck vendor menu API and exports them
to a formatted Excel file (lekki_mart_catalog.xlsx) with columns:
  SKU Code | Name | Size | Price (NGN)

Usage:
    python fetch_chowdeck_catalog.py

Requirements:
    pip install requests openpyxl
"""

import os
import requests
import openpyxl
import re
from dotenv import load_dotenv
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load environment variables
load_dotenv()

VENDOR_ID = os.getenv("CHOWDECK_VENDOR_ID")
BASE_URL = "https://api.chowdeck.com/merchant"
API_TOKEN = os.getenv("CHOWDECK_API_TOKEN")
PER_PAGE = 100
OUTPUT_FILE = "lekki_mart_catalog.xlsx"


def fetch_all_products():
    all_items = []
    headers = {"Authorization": f"Bearer {API_TOKEN}"}

    # First call - fetch first page to get total count
    params = {"per_page": PER_PAGE, "consider_variant": "true", "page": 1}
    print(f"  Fetching page 1...", end=" ")
    resp = requests.get(f"{BASE_URL}/{VENDOR_ID}/menu", params=params, headers=headers, timeout=30)
    resp.raise_for_status()
    data = resp.json()

    items = data.get("data", [])
    print(f"{len(items)} items")
    if not items:
        return []

    all_items.extend(items)

    # If we got fewer than PER_PAGE, we're done
    if len(items) < PER_PAGE:
        return all_items

    # Otherwise fetch remaining pages
    page = 2
    while page <= 100:  # Safety limit
        params = {"per_page": PER_PAGE, "consider_variant": "true", "page": page}
        print(f"  Fetching page {page}...", end=" ")
        try:
            resp = requests.get(f"{BASE_URL}/{VENDOR_ID}/menu", params=params, headers=headers, timeout=30)
            resp.raise_for_status()
            data = resp.json()
            items = data.get("data", [])
            print(f"{len(items)} items")
            if not items:
                break
            all_items.extend(items)
            if len(items) < PER_PAGE:
                break
        except Exception as e:
            print(f"Error: {e}")
            break
        page += 1

    return all_items


def extract_size(item: dict) -> str:
    size = item.get("size_description") or ""
    if size:
        return size.strip()

    name = item.get("name", "")
    match = re.search(r"\(([^)]+)\)", name)
    if match:
        candidate = match.group(1)
        if re.search(r"\d", candidate):
            return candidate.strip()

    return item.get("price_description", "").strip()


def build_sku(index: int) -> str:
    return f"LM{index:05d}"


def format_price(raw_price) -> float:
    try:
        return float(raw_price) / 100
    except (TypeError, ValueError):
        return 0.0


def write_excel(products: list):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product Catalog"

    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill("solid", start_color="1F4E79")
    ALT_FILL = PatternFill("solid", start_color="D6E4F0")
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    THIN = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    PRICE_FORMAT = '₦#,##0.00'

    headers = ["SKU Code", "Name", "Size", "Price (NGN)"]
    col_widths = [14, 48, 18, 16]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    for row_idx, item in enumerate(products, start=1):
        excel_row = row_idx + 1
        fill = ALT_FILL if row_idx % 2 == 0 else None

        sku = build_sku(row_idx)
        name = (item.get("name") or "").strip()
        size = extract_size(item)
        price = format_price(item.get("price"))

        row_data = [sku, name, size, price]
        alignments = [CENTER, LEFT, CENTER, CENTER]

        for col_idx, (value, align) in enumerate(zip(row_data, alignments), start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = align
            cell.border = BORDER
            if fill:
                cell.fill = fill

        price_cell = ws.cell(row=excel_row, column=4)
        price_cell.number_format = PRICE_FORMAT

        ws.row_dimensions[excel_row].height = 18

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{len(products) + 1}"

    wb.save(OUTPUT_FILE)
    print(f"\n Saved {OUTPUT_FILE} ({len(products)} products)")


def main():
    print(f"Fetching products for vendor {VENDOR_ID}…")
    print(f"Using token: {API_TOKEN[:20]}...")
    products = fetch_all_products()
    print(f"\nTotal products fetched: {len(products)}")

    if not products:
        print("No products found. Check vendor ID or API availability.")
        return

    print("\nBuilding Excel file...")
    write_excel(products)


if __name__ == "__main__":
    main()
    print("Script complete.")